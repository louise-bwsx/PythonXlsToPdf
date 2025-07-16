import subprocess
import os
import tempfile
import shutil
import platform
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties


def convert_xls_to_xlsx_via_libreoffice(source_path: str, xlsx_output_path: str):
    temp_dir = tempfile.mkdtemp()
    try:
        soffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice" if platform.system() == "Darwin" else "soffice"

        subprocess.run([
            soffice_path,
            "--headless",
            "--convert-to", "xlsx",
            "--outdir", temp_dir,
            source_path
        ], check=True)

        base_name = os.path.basename(source_path)
        new_xlsx_name = os.path.splitext(base_name)[0] + ".xlsx"
        converted_path = os.path.join(temp_dir, new_xlsx_name)

        shutil.move(converted_path, xlsx_output_path)
        print(f"✅ .xls 轉 .xlsx 成功：{xlsx_output_path}")
    finally:
        shutil.rmtree(temp_dir)


def estimate_display_width(text):
    width = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff':  # 中文範圍
            width += 2.5
        elif ord(ch) > 127:  # 全形字、特殊字
            width += 2
        else:
            width += 1
    return width


def normalize_excel_format(path: str):
    wb = load_workbook(path, keep_links=False, data_only=True)
    for ws in wb.worksheets:
        # ✅ 若尚未初始化 pageSetUpPr，需要補上
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties()

        # ✅ fitToPage 開啟
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # ✅ 自動欄寬調整
        for col in ws.columns:
            max_width = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        width = estimate_display_width(cell.value)
                        if width > max_width:
                            max_width = width
                except:
                    pass
            adjusted_width = max_width + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # ✅ 設定列印區與縮放
        ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1

        # ✅ 顯示格線（此設定實際轉 PDF 可能被 LibreOffice 覆蓋）
        ws.sheet_view.showGridLines = True

    wb.save(path)
    print(f"✅ 已套用列印設定與欄寬：{path}")


def convert_xlsx_to_pdf_via_libreoffice(source_path: str, output_pdf_path: str):
    # 確認 LibreOffice CLI 路徑
    soffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice" if platform.system() == "Darwin" else "soffice"

    # 建立暫存資料夾來保存轉換後的 PDF
    temp_dir = tempfile.mkdtemp()
    try:
        subprocess.run([
            soffice_path,
            "--headless",
            "--convert-to", "pdf",
            "--outdir", temp_dir,
            source_path
        ], check=True)

        # 找到轉換出的 PDF 檔名（與 Excel 同名）
        base_name = os.path.basename(source_path)
        pdf_name = os.path.splitext(base_name)[0] + ".pdf"
        generated_pdf_path = os.path.join(temp_dir, pdf_name)

        # 搬到目標路徑
        shutil.move(generated_pdf_path, output_pdf_path)
        print(f"✅ 已成功轉出 PDF：{output_pdf_path}")
    except Exception as e:
        print("❌ 轉換失敗：", e)
    finally:
        shutil.rmtree(temp_dir)

# 測試時取消註解
# # 將xls轉成xlsx
# convert_xls_to_xlsx_via_libreoffice("DemoChartZJPGMerge100檔案.xls", "temp.xlsx")
# # 格式修正
# normalize_excel_format("temp.xlsx")
# # 執行範例
# convert_xlsx_to_pdf_via_libreoffice("temp.xlsx", "output.pdf")

# 給API呼叫時取消註解
if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("請提供參數：<輸入 xls 路徑> <輸出 pdf 路徑>")
        sys.exit(1)
    input_xls = sys.argv[1]
    output_pdf = sys.argv[2]
    # 中繼轉成 .xlsx
    temp_xlsx = os.path.join(tempfile.gettempdir(
    ), "converted_" + os.path.basename(input_xls).replace(".xls", ".xlsx"))
    convert_xls_to_xlsx_via_libreoffice(input_xls, temp_xlsx)
    normalize_excel_format(temp_xlsx)
    convert_xlsx_to_pdf_via_libreoffice(temp_xlsx, output_pdf)
